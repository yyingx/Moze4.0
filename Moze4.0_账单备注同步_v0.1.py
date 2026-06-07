# -*- coding: utf-8 -*-
"""
Moze 4.0 账单备注同步工具。

使用流程：
1. 弹窗选择任意带日期区间的vivo账单。
2. 从vivo账单文件名解析日期区间。
3. 自动识别相同日期区间的原始账单：
   - 微信支付账单流水文件(YYYYMMDD-YYYYMMDD)_*.xlsx
   - 支付宝交易明细(YYYYMMDD-YYYYMMDD).csv
4. 按金额和交易时间匹配记录，默认只填充原始账单里空白或 "/" 的“备注”列。
   如需覆盖已有备注，运行时添加 --overwrite-existing。

备注提取不强依赖空格。脚本会先匹配到原始账单记录，再从vivo账单备注里剥离原始账单的
交易对方、商品说明和脱敏括号，剩余内容作为有效备注。手写的 "|" 或 "｜"
仍作为显式分隔符优先支持。
"""

from __future__ import annotations

import argparse
import csv
import re
import shutil
import sys
import tkinter as tk
import warnings
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from tkinter import filedialog

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


PROGRAM_NAME = "Moze4.0_账单备注同步"
__version__ = "0.1"
__updated__ = "2026-06-07"

CURRENT_DIR = Path(__file__).resolve().parent if "__file__" in globals() else Path.cwd()
TIME_WINDOW_SECONDS = 3

SOURCE_REQUIRED_HEADERS = ("账单日期", "备注", "金额")
WECHAT_REQUIRED_HEADERS = ("交易时间", "交易对方", "商品", "金额(元)", "备注")
ALIPAY_REQUIRED_HEADERS = ("交易时间", "交易对方", "商品说明", "金额", "备注")

SEPARATOR_RE = re.compile(r"[|｜]", re.ASCII)
SPACE_NOTE_RE = re.compile(r"^\s*(?P<prefix>.*?)\s+(?P<note>.+?)\s*$")
PRIVACY_BRACKET_RE = re.compile(r"[（(][^）)]*[*＊][^）)]*[）)]")
GENERIC_PRODUCT_PARTS = {"", "/", "消费", "二维码收款", "收款方备注", "自助服务"}


@dataclass(frozen=True)
class FileDateRange:
    start: date
    end: date


@dataclass(frozen=True)
class SourceRow:
    row: int
    dt: datetime
    amount: Decimal
    remark: str


@dataclass(frozen=True)
class BillRow:
    row_id: int
    display_row: int
    dt: datetime
    amount: Decimal
    party: str
    product: str
    remark: str


@dataclass(frozen=True)
class Match:
    source: SourceRow
    target: BillRow
    note: str
    diff_seconds: float
    score: int


@dataclass
class BillResult:
    label: str
    path: Path
    matches: list[Match]
    no_note: list[str]
    problems: list[str]
    unmatched_targets: int
    skipped_existing: int
    row_count: int
    header_row: int


def fmt_dt(value: datetime) -> str:
    if value.microsecond == 999999:
        return value.strftime("%Y-%m-%d %H:%M")
    return value.strftime("%Y-%m-%d %H:%M:%S")


def fmt_amount(value: Decimal) -> str:
    return f"{value.normalize():f}"


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_text(value: Any) -> str:
    return re.sub(r"\s+", "", cell_text(value))


def normalize_header_name(value: Any) -> str:
    return (
        cell_text(value)
        .replace("（", "(")
        .replace("）", ")")
        .replace("\ufeff", "")
    )


def parse_datetime(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())

    text = cell_text(value)
    if not text:
        return None

    formats = (
        "%Y-%m-%d %H:%M:%S.%f",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y/%m/%d %H:%M:%S.%f",
        "%Y/%m/%d %H:%M:%S",
        "%Y/%m/%d %H:%M",
    )
    for fmt in formats:
        try:
            parsed = datetime.strptime(text, fmt)
            if fmt.endswith("%H:%M"):
                return parsed.replace(microsecond=999999)
            return parsed
        except ValueError:
            pass
    return None


def parse_amount(value: Any) -> Decimal | None:
    text = cell_text(value)
    if not text:
        return None
    text = text.replace(",", "").replace("¥", "").replace("￥", "")
    text = text.replace("元", "").strip()
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def parse_yyyymmdd(value: str) -> date:
    return datetime.strptime(value, "%Y%m%d").date()


def parse_dotted_date(year: str, month: str, day: str) -> date:
    return date(int(year), int(month), int(day))


def parse_range_from_filename(path: Path) -> FileDateRange | None:
    name = path.name

    compact = re.search(r"\((\d{8})-(\d{8})\)", name)
    if compact:
        return FileDateRange(parse_yyyymmdd(compact.group(1)), parse_yyyymmdd(compact.group(2)))

    dotted = re.search(
        r"(\d{4})\.(\d{2})\.(\d{2})-(\d{4})\.(\d{2})\.(\d{2})",
        name,
    )
    if dotted:
        return FileDateRange(
            parse_dotted_date(dotted.group(1), dotted.group(2), dotted.group(3)),
            parse_dotted_date(dotted.group(4), dotted.group(5), dotted.group(6)),
        )

    return None


def compact_range(date_range: FileDateRange) -> str:
    return f"{date_range.start:%Y%m%d}-{date_range.end:%Y%m%d}"


def dotted_range(date_range: FileDateRange) -> str:
    return f"{date_range.start:%Y.%m.%d}-{date_range.end:%Y.%m.%d}"


def is_blank_xlsx_row(ws: Worksheet, row: int) -> bool:
    return all(cell_text(ws.cell(row=row, column=col).value) == "" for col in range(1, ws.max_column + 1))


def find_xlsx_header_row(
    ws: Worksheet,
    required_headers: tuple[str, ...],
    max_scan_rows: int = 100,
) -> tuple[int, dict[str, int]]:
    max_row = min(ws.max_row, max_scan_rows)
    for row in range(1, max_row + 1):
        header_map = {
            normalize_header_name(ws.cell(row=row, column=col).value): col
            for col in range(1, ws.max_column + 1)
            if normalize_header_name(ws.cell(row=row, column=col).value)
        }
        if all(header in header_map for header in required_headers):
            return row, header_map
    raise ValueError(f"未找到表头: {', '.join(required_headers)}")


def find_csv_header_row(rows: list[list[str]], required_headers: tuple[str, ...]) -> tuple[int, dict[str, int]]:
    for index, row in enumerate(rows):
        header_map = {normalize_header_name(value): col for col, value in enumerate(row) if normalize_header_name(value)}
        if all(header in header_map for header in required_headers):
            return index, header_map
    raise ValueError(f"CSV 未找到表头: {', '.join(required_headers)}")


def resolve_path(path_text: str, base_dir: Path) -> Path:
    path = Path(path_text)
    if not path.is_absolute():
        path = base_dir / path
    return path


def source_sort_key(path: Path) -> tuple[date, date, float]:
    date_range = parse_range_from_filename(path)
    if date_range is None:
        return date.min, date.min, path.stat().st_mtime
    return date_range.start, date_range.end, path.stat().st_mtime


def display_candidate(path: Path) -> str:
    date_range = parse_range_from_filename(path)
    if date_range is None:
        return path.name
    return f"{dotted_range(date_range)}  {path.name}"


def choose_from_list(prompt: str, candidates: list[Path], base_dir: Path) -> Path:
    print(prompt)
    for index, path in enumerate(candidates, start=1):
        print(f"  {index}. {display_candidate(path)}")

    while True:
        choice = input("请选择序号、文件名或完整路径，直接回车选择 1: ").strip().strip('"').strip("'")
        if not choice:
            return candidates[0]
        if choice.isdigit():
            selected = int(choice)
            if 1 <= selected <= len(candidates):
                return candidates[selected - 1]

        named_candidate = next((path for path in candidates if path.name == choice), None)
        if named_candidate is not None:
            return named_candidate

        manual_path = Path(choice)
        if not manual_path.is_absolute():
            manual_path = base_dir / manual_path
        if manual_path.exists():
            return manual_path

        print("输入无效，请重新选择。")


def choose_source_with_dialog(base_dir: Path) -> Path | None:
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    root.lift()
    root.focus_force()
    root.update()

    print("请在弹出的窗口中选择vivo账单 xlsx 文件...")
    selected = filedialog.askopenfilename(
        parent=root,
        title="请选择vivo账单文件",
        initialdir=str(base_dir),
        filetypes=[
            ("vivo账单 Excel", "*.xlsx"),
            ("所有文件", "*.*"),
        ],
    )
    root.attributes("-topmost", False)
    root.update()
    root.destroy()

    if not selected:
        return None
    return Path(selected)


def choose_source_from_terminal(base_dir: Path) -> Path:
    candidates = [
        path
        for path in base_dir.glob("*.xlsx")
        if parse_range_from_filename(path) is not None
        and not path.name.startswith("微信支付账单流水文件")
        and not path.name.startswith("~$")
    ]
    candidates.sort(key=source_sort_key, reverse=True)
    if not candidates:
        raise FileNotFoundError("当前目录未找到带日期区间的vivo账单 xlsx 文件。")
    print(f"扫描目录: {base_dir}")
    print(f"找到vivo账单候选: {len(candidates)} 个")
    return choose_from_list("请选择vivo账单文件：", candidates, base_dir)


def choose_source_interactively(base_dir: Path) -> Path:
    selected = choose_source_with_dialog(base_dir)
    if selected is not None:
        return selected

    print("未选择文件，改用终端候选列表。")
    return choose_source_from_terminal(base_dir)


def pick_candidate(label: str, candidates: list[Path], interactive: bool, base_dir: Path) -> Path | None:
    if not candidates:
        return None
    candidates = sorted(candidates, key=lambda path: path.stat().st_mtime, reverse=True)
    if len(candidates) == 1 or not interactive:
        print(f"自动识别{label}: {candidates[0].name}")
        return candidates[0]
    return choose_from_list(f"发现多个{label}候选，请选择：", candidates, base_dir)


def discover_targets(source_path: Path, base_dir: Path, interactive: bool) -> tuple[Path | None, Path | None]:
    date_range = parse_range_from_filename(source_path)
    if date_range is None:
        raise ValueError(f"无法从vivo账单文件名识别日期区间: {source_path.name}")

    compact = compact_range(date_range)
    print(f"已选择vivo账单日期区间: {dotted_range(date_range)} ({compact})")
    wechat_candidates = list(base_dir.glob(f"微信支付账单流水文件({compact})*.xlsx"))
    alipay_candidates = list(base_dir.glob(f"支付宝交易明细({compact})*.csv"))

    wechat_path = pick_candidate("微信账单", wechat_candidates, interactive, base_dir)
    alipay_path = pick_candidate("支付宝账单", alipay_candidates, interactive, base_dir)
    return wechat_path, alipay_path


def validate_source_path(source_path: Path) -> None:
    if source_path.suffix.lower() != ".xlsx":
        raise ValueError(f"vivo账单必须是 xlsx 文件: {source_path.name}")
    if source_path.name.startswith("微信支付账单流水文件"):
        raise ValueError(f"请选择vivo账单，不要选择微信账单: {source_path.name}")
    if source_path.name.startswith("~$"):
        raise ValueError(f"不能选择 Excel 临时文件: {source_path.name}")
    if parse_range_from_filename(source_path) is None:
        raise ValueError(f"vivo账单文件名未包含日期区间: {source_path.name}")


def load_source_rows(ws: Worksheet) -> tuple[list[SourceRow], int]:
    header_row, headers = find_xlsx_header_row(ws, SOURCE_REQUIRED_HEADERS)
    date_col = headers["账单日期"]
    remark_col = headers["备注"]
    amount_col = headers["金额"]

    rows: list[SourceRow] = []
    skipped_invalid = 0

    for row in range(header_row + 1, ws.max_row + 1):
        if is_blank_xlsx_row(ws, row):
            continue

        dt = parse_datetime(ws.cell(row=row, column=date_col).value)
        amount = parse_amount(ws.cell(row=row, column=amount_col).value)
        remark = cell_text(ws.cell(row=row, column=remark_col).value)
        if dt is None or amount is None or not remark:
            skipped_invalid += 1
            continue

        rows.append(SourceRow(row=row, dt=dt, amount=amount, remark=remark))

    return rows, skipped_invalid


def load_wechat_rows(ws: Worksheet) -> tuple[list[BillRow], dict[str, int], int]:
    header_row, headers = find_xlsx_header_row(ws, WECHAT_REQUIRED_HEADERS)
    time_col = headers["交易时间"]
    party_col = headers["交易对方"]
    product_col = headers["商品"]
    amount_col = headers["金额(元)"]
    remark_col = headers["备注"]

    rows: list[BillRow] = []
    for row in range(header_row + 1, ws.max_row + 1):
        if is_blank_xlsx_row(ws, row):
            break

        dt = parse_datetime(ws.cell(row=row, column=time_col).value)
        amount = parse_amount(ws.cell(row=row, column=amount_col).value)
        if dt is None or amount is None:
            continue

        rows.append(
            BillRow(
                row_id=row,
                display_row=row,
                dt=dt,
                amount=amount,
                party=cell_text(ws.cell(row=row, column=party_col).value),
                product=cell_text(ws.cell(row=row, column=product_col).value),
                remark=cell_text(ws.cell(row=row, column=remark_col).value),
            )
        )

    return rows, headers, header_row


def detect_csv_encoding(path: Path) -> str:
    data = path.read_bytes()
    for encoding in ("utf-8-sig", "gb18030", "gbk", "utf-8"):
        try:
            data.decode(encoding)
            return encoding
        except UnicodeDecodeError:
            continue
    return "gb18030"


def detect_line_terminator(path: Path) -> str:
    data = path.read_bytes()
    return "\r\n" if b"\r\n" in data else "\n"


def read_csv_rows(path: Path, encoding: str) -> list[list[str]]:
    with path.open("r", encoding=encoding, newline="") as handle:
        return list(csv.reader(handle))


def load_alipay_rows(rows: list[list[str]]) -> tuple[list[BillRow], dict[str, int], int]:
    header_index, headers = find_csv_header_row(rows, ALIPAY_REQUIRED_HEADERS)
    time_col = headers["交易时间"]
    party_col = headers["交易对方"]
    product_col = headers["商品说明"]
    amount_col = headers["金额"]
    remark_col = headers["备注"]

    bill_rows: list[BillRow] = []
    for index in range(header_index + 1, len(rows)):
        row = rows[index]
        if not row or all(cell_text(value) == "" for value in row):
            continue

        def get(col: int) -> str:
            return row[col] if col < len(row) else ""

        dt = parse_datetime(get(time_col))
        amount = parse_amount(get(amount_col))
        if dt is None or amount is None:
            continue

        bill_rows.append(
            BillRow(
                row_id=index,
                display_row=index + 1,
                dt=dt,
                amount=amount,
                party=cell_text(get(party_col)),
                product=cell_text(get(product_col)),
                remark=cell_text(get(remark_col)),
            )
        )

    return bill_rows, headers, header_index + 1


def contains_regex_safe(needle: str, haystack: str) -> bool:
    needle = normalize_text(needle)
    haystack = normalize_text(haystack)
    if not needle or needle == "/":
        return False
    return re.search(re.escape(needle), haystack) is not None


def product_parts(product: str) -> list[str]:
    product = cell_text(product)
    if not product or product == "/":
        return []

    product = re.sub(r"^收款方备注[:：]", "", product).strip()
    if not product or "二维码收款" in product:
        return []

    parts = [product]
    parts.extend(re.split(r"[-/／]", product))
    cleaned: list[str] = []
    for part in parts:
        part = cell_text(part)
        if part in GENERIC_PRODUCT_PARTS:
            continue
        cleaned.append(part)
    return cleaned


def target_names(target: BillRow) -> list[str]:
    names = [target.party]
    names.extend(product_parts(target.product))

    unique: list[str] = []
    seen: set[str] = set()
    for name in sorted(names, key=lambda item: len(item), reverse=True):
        key = normalize_text(name)
        if not key or key in seen or key in GENERIC_PRODUCT_PARTS:
            continue
        seen.add(key)
        unique.append(name)
    return unique


def score_candidate(source: SourceRow, target: BillRow) -> int:
    score = 0
    if contains_regex_safe(target.party, source.remark):
        score += 5

    for part in product_parts(target.product):
        if contains_regex_safe(part, source.remark):
            score += 2
            break

    return score


def interval_for_time(value: datetime) -> tuple[datetime, datetime]:
    if value.microsecond == 999999:
        start = value.replace(second=0, microsecond=0)
        return start, start + timedelta(seconds=59, microseconds=999999)
    if value.microsecond == 0:
        return value, value + timedelta(microseconds=999999)
    return value, value


def fuzzy_time_diff_seconds(left: datetime, right: datetime) -> float:
    left_start, left_end = interval_for_time(left)
    right_start, right_end = interval_for_time(right)

    if left_start <= right_end and right_start <= left_end:
        return 0.0
    if left_end < right_start:
        return (right_start - left_end).total_seconds()
    return (left_start - right_end).total_seconds()


def strip_note_noise(value: str) -> str:
    value = PRIVACY_BRACKET_RE.sub("", value)
    value = re.sub(r"\s+", " ", value)
    return value.strip(" \t\r\n-—–_,，.。:：;；/\\|｜")


def remove_first_occurrence(text: str, needle: str) -> tuple[str, bool]:
    needle = cell_text(needle)
    if not needle or needle == "/":
        return text, False
    updated, count = re.subn(re.escape(needle), "", text, count=1)
    return updated, count > 0


def normalize_note(note: str, source: SourceRow, target: BillRow) -> str | None:
    note = strip_note_noise(note)
    if not note or note == "/":
        return None
    if normalize_text(note) == normalize_text(source.remark):
        return None

    blocked = {normalize_text(target.party), normalize_text(target.product)}
    blocked.update(normalize_text(part) for part in product_parts(target.product))
    if normalize_text(note) in blocked:
        return None

    return note


def bill_summary(target: BillRow) -> str:
    parts = [fmt_dt(target.dt)]
    if target.party:
        parts.append(target.party)
    parts.append(fmt_amount(target.amount))
    return " ".join(parts)


def source_summary(source: SourceRow) -> str:
    return f"{fmt_dt(source.dt)} {source.remark} {fmt_amount(source.amount)}"


def extract_by_separator(remark: str) -> str | None:
    parts = SEPARATOR_RE.split(remark, maxsplit=1)
    if len(parts) != 2:
        return None
    note = strip_note_noise(parts[1])
    return note or None


def extract_by_space_fallback(source: SourceRow, target: BillRow) -> str | None:
    match = SPACE_NOTE_RE.match(source.remark)
    if not match:
        return None

    prefix = match.group("prefix")
    note = match.group("note")
    if not contains_regex_safe(target.party, prefix) and not any(
        contains_regex_safe(part, prefix) for part in product_parts(target.product)
    ):
        return None

    return normalize_note(note, source, target)


def extract_effective_note(source: SourceRow, target: BillRow) -> str | None:
    explicit = extract_by_separator(source.remark)
    if explicit:
        return explicit

    text = source.remark
    removed_any = False
    for name in target_names(target):
        text, removed = remove_first_occurrence(text, name)
        removed_any = removed_any or removed

    if removed_any:
        note = normalize_note(text, source, target)
        if note:
            return note

    return extract_by_space_fallback(source, target)


def build_matches(
    source_rows: list[SourceRow],
    target_rows: list[BillRow],
    time_window_seconds: float,
    only_empty: bool,
) -> tuple[list[Match], list[str], list[str], int, int]:
    matches: list[Match] = []
    no_note: list[str] = []
    problems: list[str] = []
    unmatched_targets = 0
    skipped_existing = 0
    used_source_rows: set[int] = set()

    for target in target_rows:
        if only_empty and target.remark not in ("", "/"):
            skipped_existing += 1
            continue

        candidates: list[tuple[int, float, SourceRow]] = []
        for source in source_rows:
            if source.row in used_source_rows:
                continue
            if source.amount != target.amount:
                continue

            diff_seconds = fuzzy_time_diff_seconds(source.dt, target.dt)
            if diff_seconds <= time_window_seconds:
                candidates.append((score_candidate(source, target), diff_seconds, source))

        if not candidates:
            unmatched_targets += 1
            continue

        candidates.sort(key=lambda item: (-item[0], item[1], item[2].row))
        best_score, best_diff, best_source = candidates[0]
        tied = [item for item in candidates if item[0] == best_score and item[1] == best_diff]

        if len(tied) > 1:
            source_rows_text = "；".join(source_summary(item[2]) for item in tied)
            problems.append(f"冲突跳过: {bill_summary(target)} 可匹配 {source_rows_text}")
            continue

        used_source_rows.add(best_source.row)
        note = extract_effective_note(best_source, target)
        if not note:
            no_note.append(f"{bill_summary(target)} <- {source_summary(best_source)}，无有效备注")
            continue

        matches.append(Match(source=best_source, target=target, note=note, diff_seconds=best_diff, score=best_score))

    return matches, no_note, problems, unmatched_targets, skipped_existing


def print_bill_result(result: BillResult) -> None:
    print(f"\n{result.label}: {result.path.name}")
    print(f"账单流水: {result.row_count} 行")
    print(f"匹配到可写备注: {len(result.matches)} 行")

    for match in result.matches:
        old = match.target.remark or ""
        print(
            f"  {bill_summary(match.target)} "
            f"时间差={match.diff_seconds:.3f}s 备注: {old!r} -> {match.note!r}"
        )

    if result.no_note:
        print(f"匹配但无有效备注: {len(result.no_note)} 行")
        for item in result.no_note:
            print(f"  {item}")

    if result.unmatched_targets:
        print(f"账单未在vivo账单找到对应记录: {result.unmatched_targets} 行")

    if result.skipped_existing:
        print(f"已有备注跳过: {result.skipped_existing} 行")

    if result.problems:
        print(f"未写入/需检查: {len(result.problems)} 行")
        for problem in result.problems:
            print(f"  {problem}")


def backup_path_for(path: Path) -> Path:
    return path.with_name(f"{path.stem}.bak{path.suffix}")


def process_wechat(
    path: Path,
    source_rows: list[SourceRow],
    args: argparse.Namespace,
) -> tuple[BillResult, bool]:
    wb = load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    target_rows, headers, header_row = load_wechat_rows(ws)
    matches, no_note, problems, unmatched, skipped_existing = build_matches(
        source_rows,
        target_rows,
        args.time_window,
        only_empty=not args.overwrite_existing,
    )

    result = BillResult(
        label="微信账单",
        path=path,
        matches=matches,
        no_note=no_note,
        problems=problems,
        unmatched_targets=unmatched,
        skipped_existing=skipped_existing,
        row_count=len(target_rows),
        header_row=header_row,
    )

    if args.dry_run or not matches:
        return result, True

    remark_col = headers["备注"]
    for match in matches:
        ws.cell(row=match.target.row_id, column=remark_col).value = match.note

    if args.backup:
        backup_path = backup_path_for(path)
        shutil.copy2(path, backup_path)
        print(f"已备份: {backup_path}")

    try:
        wb.save(path)
    except PermissionError:
        print(f"保存失败: 无法写入 {path}。请先关闭 Excel/WPS 中打开的这个文件。", file=sys.stderr)
        return result, False

    print(f"已覆盖保存: {path}")
    return result, True


def process_alipay(
    path: Path,
    source_rows: list[SourceRow],
    args: argparse.Namespace,
) -> tuple[BillResult, bool]:
    encoding = detect_csv_encoding(path)
    line_terminator = detect_line_terminator(path)
    rows = read_csv_rows(path, encoding)
    target_rows, headers, header_row = load_alipay_rows(rows)
    matches, no_note, problems, unmatched, skipped_existing = build_matches(
        source_rows,
        target_rows,
        args.time_window,
        only_empty=not args.overwrite_existing,
    )

    result = BillResult(
        label="支付宝账单",
        path=path,
        matches=matches,
        no_note=no_note,
        problems=problems,
        unmatched_targets=unmatched,
        skipped_existing=skipped_existing,
        row_count=len(target_rows),
        header_row=header_row,
    )

    if args.dry_run or not matches:
        return result, True

    remark_col = headers["备注"]
    for match in matches:
        row = rows[match.target.row_id]
        while len(row) <= remark_col:
            row.append("")
        row[remark_col] = match.note

    if args.backup:
        backup_path = backup_path_for(path)
        shutil.copy2(path, backup_path)
        print(f"已备份: {backup_path}")

    try:
        with path.open("w", encoding=encoding, newline="") as handle:
            writer = csv.writer(handle, lineterminator=line_terminator)
            writer.writerows(rows)
    except PermissionError:
        print(f"保存失败: 无法写入 {path}。请先关闭 Excel/WPS 中打开的这个文件。", file=sys.stderr)
        return result, False

    print(f"已覆盖保存: {path}")
    return result, True


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="选择vivo账单后，自动把有效备注同步到同日期区间的微信/支付宝账单。")
    parser.add_argument("--version", action="version", version=f"{PROGRAM_NAME} v{__version__}")
    parser.add_argument("--source", help="vivo账单路径；不传则弹窗选择")
    parser.add_argument("--wechat", help="微信账单路径；不传则按vivo账单日期区间自动识别")
    parser.add_argument("--alipay", help="支付宝账单路径；不传则按vivo账单日期区间自动识别")
    parser.add_argument("--time-window", type=float, default=TIME_WINDOW_SECONDS, help="允许的时间差秒数，默认3")
    parser.add_argument("--dry-run", action="store_true", help="只打印匹配结果，不保存")
    parser.add_argument("--overwrite-existing", action="store_true", help="允许覆盖微信/支付宝账单中已有备注")
    parser.add_argument("--backup", action="store_true", help="保存前生成 .bak 文件")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv or sys.argv[1:])
    base_dir = CURRENT_DIR

    try:
        source_path = resolve_path(args.source, base_dir) if args.source else choose_source_interactively(base_dir)
        validate_source_path(source_path)
        source_dir = source_path.parent

        if args.wechat or args.alipay:
            wechat_path = resolve_path(args.wechat, source_dir) if args.wechat else None
            alipay_path = resolve_path(args.alipay, source_dir) if args.alipay else None
        else:
            wechat_path, alipay_path = discover_targets(source_path, source_dir, interactive=not args.source)
    except (FileNotFoundError, ValueError) as exc:
        print(str(exc), file=sys.stderr)
        return 1

    if not source_path.exists():
        print(f"vivo账单不存在: {source_path}", file=sys.stderr)
        return 1

    if wechat_path is None and alipay_path is None:
        print("未找到同日期区间的微信或支付宝账单。", file=sys.stderr)
        return 1

    warnings.filterwarnings("ignore", message="Workbook contains no default style")

    source_wb = load_workbook(source_path, read_only=True, data_only=True)
    source_ws = source_wb[source_wb.sheetnames[0]]
    source_rows, source_skipped_invalid = load_source_rows(source_ws)

    print(f"vivo账单: {source_path.name}")
    print(f"vivo账单有效数据: {len(source_rows)} 行，无效/空行跳过: {source_skipped_invalid} 行")

    success = True
    if wechat_path is not None:
        if not wechat_path.exists():
            print(f"微信账单不存在: {wechat_path}", file=sys.stderr)
            success = False
        else:
            result, saved = process_wechat(wechat_path, source_rows, args)
            print_bill_result(result)
            success = success and saved

    if alipay_path is not None:
        if not alipay_path.exists():
            print(f"支付宝账单不存在: {alipay_path}", file=sys.stderr)
            success = False
        else:
            result, saved = process_alipay(alipay_path, source_rows, args)
            print_bill_result(result)
            success = success and saved

    if args.dry_run:
        print("\ndry-run：未保存文件。")

    return 0 if success else 2


if __name__ == "__main__":
    raise SystemExit(main())
